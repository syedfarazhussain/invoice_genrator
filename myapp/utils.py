import os, shutil, csv, io, traceback, sys, xlsxwriter, json, datetime, logging, pythoncom
from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse
from account.models import User
from time import sleep
import pandas as pd
from .models import AccountList, CpDeskInvoices, MasterData, Invoices, EmailCondition, GroupComp
from PyPDF2 import PdfMerger
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from docxtpl import DocxTemplate
from docx2pdf import convert
from django.conf import settings


from .sendEmail import send_mail, send_consolidated_mail, send_group_mail



output_dir = settings.MEDIA_ROOT[1]
archive_dir = settings.MEDIA_ROOT[2]
input_dir = settings.MEDIA_ROOT[3]

if settings.EMAIL:
    ConsolidatedEmail = settings.EMAIL[0]['ConsolidatedEmail']
    isSendEmail = settings.EMAIL[0]['isSendEmail']
else:
    pass

def get_num_invoice(previous_month,previous_year,current_year,account_name,type):
    
    is_invoice_exist = False

    invoices = Invoices.objects.filter(
                                account_name = account_name, 
                                previous_month = previous_month,
                                previous_year = previous_year,
                                invoice_Type = type ).order_by('-id')
    if invoices:
        is_invoice_exist = True

    if is_invoice_exist:
        invoice_number = invoices.first().invoice_number
    else:
        invoices = Invoices.objects.filter(
            current_year=current_year
        ).order_by('-id')

        if invoices:
            invoice_number = invoices.first().invoice_number
        else:
            invoice_number = f'{current_year}000000'

    return int(invoice_number)

def check_email_condition(account_name):
    email_condition = EmailCondition.objects.filter(contact_name=account_name).first()

    if email_condition is not None and email_condition.exists():
        obj = email_condition.first()
        pass
    else:
        if email_condition is not None and hasattr(email_condition, 'type') and email_condition.type.lower() == 'auto':
            return True
    
    return False

def check_group_list(account_name):
    group_comp = GroupComp.objects.filter(Child_Comp=account_name).first()

    if group_comp is not None and group_comp.exists():
        return False
    else:
        return group_comp

def process_de_data(request, lift_fee_json_data, user_data):
    Word_To_PDF_OBJ = None
    current_user = User.objects.filter(id=user_data['user']).first()
    try:
        data = json.loads(lift_fee_json_data)
        df = pd.DataFrame(data)
        with io.BytesIO() as output:
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False)
            workbook = output.getvalue()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=output.xlsx'
        response.write(workbook)
    except Exception as e:
        messages.error(request, f"!Error Occurred: {str(e)}")
    messages.info(request, '!Checking File')

    processed_data = {}
    df = pd.read_excel(workbook)

    for index, row in df.iterrows():
        if pd.isna(row['AccountName']) == False:
            if row['AccountName'] in processed_data:
                processed_data[row['AccountName']].append(row.to_dict())
            else:
                processed_data[row['AccountName']] = [row.to_dict()]

    messages.info(request, '!Generating invoices')
    Word_To_PDF_OBJ = Word_To_PDF()
    
    isSendEmail = 'True'
    type = 'normal'
    today = datetime.date.today()
    current_year_number = today.strftime("%Y")
    last_month_number = str(int(today.strftime("%m")) - 1).zfill(2)
    last_year_number = today.strftime("%Y")
    current_date_format = today.strftime("%d%m%Y")
    Consolidated_Report = f'Consolidated_Report_{current_date_format}.csv'
    dic_group_comp = {}
    confirmation_exists = 'False'

    with open(f"{output_dir}/{Consolidated_Report}", 'w', newline='') as consolidated_write_file:
        consolidated_writer = csv.writer(consolidated_write_file, csv.excel)
        consolidated_writer.writerow(['Account Name', 'Email_ids', 'Invoice_Number', 'VAT Type', 'Total USD Amount', 'GBP Rate','User'])

        for account_name in processed_data:
            total_amount = 0
            group_parent_list = []
            is_considered = AccountList.objects.filter(account_name=account_name)
            try:
                if is_considered:
                    for row in processed_data[account_name]:
                        total_amount += row['USD AMOUNT']
                    total_amount = str(round(float(total_amount), 2))

                    master_data = MasterData.objects.filter(contact_name=account_name)
                    

                    if len(master_data) == 0:
                        try:
                            consolidated_writer.writerow([account_name, '', 'Account does not exist in MasterData', '', str(total_amount),
                                                        str(user_data['gbp_rate']), current_user.username])
                        except Exception as e:
                            print(f"An error occurred while writing the row to the CSV file: {e}")
                        pass
                    else:
                        master_data = master_data[0]

                        if master_data.status.lower() == "active":
                            customer_address = master_data.customer_address
                            vat_type = master_data.vat_type
                            email_ids_TO = master_data.email_id.split(";")
                            email_ids_CC = master_data.cc_email_id.split(";")
                            cleanedList = [x for x in email_ids_CC if str(x) != 'nan']
                            email_ids = email_ids_TO + cleanedList


                            invoice_exists = Invoices.objects.filter(account_name=account_name, previous_month=last_month_number, previous_year=last_year_number, invoice_Type=type).exists()


                            confirm_creating_invoice = True

                            if not invoice_exists == False:
                                if confirmation_exists.lower() == 'false':
                                    confirm_creating_invoice = True          # user input required yes or no to re generate invoice 
                                    if confirm_creating_invoice:
                                        confirmation_exists = 'true'

                            num_invoice = get_num_invoice(last_month_number, last_year_number, current_year_number, account_name, type)
                            if invoice_exists:
                                num_invoice = int(num_invoice) - 1

                            if confirm_creating_invoice:
                                messages.success(request, f"!Creating invoice for {account_name}")

                                create_invoice = Word_To_PDF_OBJ.generate_pdf(account_name, customer_address, str(total_amount), vat_type, user_data['gbp_rate'], str(num_invoice + 1))
                                invoice_pdf_file = create_invoice[0]
                                invoice_number = create_invoice[1]
                                invoice_excel_file = Word_To_PDF_OBJ.genrate_excel(account_name, processed_data[account_name])
                                files = [f'{output_dir}/{invoice_pdf_file}', f'{output_dir}/{invoice_excel_file}']

                                messages.info(request, f"!Sending email to {account_name}")

                                is_email_auto = check_email_condition(account_name)

                                if is_email_auto:
                                    please_confirm_text = "We will deduct these lifting fees from the respective currency balance."
                                else:
                                    please_confirm_text = "Please confirm we are able to deduct these lifting fees from the respective currency balance."

                                is_exist_in_group = check_group_list(account_name)
                                
                                if not is_exist_in_group:
                                        if isSendEmail.lower() == 'true':
                                            send_mail(
                                                send_from=user_data['smtp_sender'],
                                                send_to=email_ids_TO,
                                                send_cc=cleanedList,
                                                files=files,
                                                invoice_number=invoice_number,
                                                receiver_name=account_name,
                                                total_amount=total_amount,
                                                smtp_server=user_data['smtp_server'],
                                                smtp_port=user_data['smtp_port'],
                                                smtp_user=user_data['smtp_user'],
                                                smtp_pass=user_data['smtp_pass'],
                                                please_confirm_text=please_confirm_text
                                            )
                                            logging.basicConfig(
                                                filename=f"{archive_dir}/Logger_{current_date_format}.log",
                                                level=logging.INFO)
                                            logging.info(str(account_name) + str(str(email_ids_TO) + str(cleanedList)) + str(files) )

                                            messages.success(request, f"!Email sent to {account_name}")

                                        shutil.move(
                                            f'{output_dir}/{invoice_pdf_file}',
                                            f'{archive_dir}/{invoice_pdf_file}'
                                        )
                                        shutil.move(
                                            f'{output_dir}/{invoice_excel_file}',
                                            f'{archive_dir}/{invoice_excel_file}'
                                        )
                                       
                                        try:
                                             consolidated_writer.writerow(
                                                    [account_name, email_ids, str(num_invoice + 1), vat_type, str(total_amount),
                                                    str(user_data['gbp_rate']), current_user.username])
                                        except Exception as e:
                                            print(f"An error occurred while writing the row to the CSV file: {e}")
                                       
                                else:
                                    group_comp_parent = is_exist_in_group[2]
                                    parent_comp_name = is_exist_in_group[3]

                                    dicAcc = {}
                                    listChild = []
                                    dicAcc['account_name'] = account_name
                                    dicAcc['pdf_file'] = f'{output_dir}/{invoice_pdf_file}'
                                    dicAcc['excel_file'] = f'{output_dir}/{invoice_excel_file}'
                                    dicAcc['email_ids_TO'] = email_ids_TO
                                    dicAcc['email_ids_cc'] = cleanedList
                                    dicAcc['parent_comp_name'] = parent_comp_name
                                    if group_comp_parent in dic_group_comp.keys():
                                        listChild = dic_group_comp[group_comp_parent]
                                        listChild.append(dicAcc)
                                        dic_group_comp[group_comp_parent] = listChild
                                    else:
                                        listChild.append(dicAcc)
                                        dic_group_comp[group_comp_parent] = listChild

                                    consolidated_writer.writerow(
                                        [account_name, email_ids, 'Clubbed in Group Invoice', vat_type,
                                            str(total_amount),
                                            str(user_data['gbp_rate']), current_user.username])
                                    
                                if not invoice_exists:
                                    if not invoice_pdf_file.endswith(".pdf"):
                                        invoice_pdf_file += ".pdf"
                                    invoices = Invoices.objects.create(
                                        account_name=account_name, 
                                        invoice_file_name=invoice_pdf_file,
                                        invoice_number=invoice_number, 
                                        previous_month=last_month_number,
                                        previous_year=last_year_number,
                                        current_year=current_year_number,
                                        total_cost=str(total_amount), 
                                        user_name=current_user.username,
                                        invoice_Type=type
                                    )
                                    invoices.user = request.user
                                    invoices.save()
                                return invoice_pdf_file
                        else:
                            consolidated_writer.writerow(
                                [account_name, '', 'Account is not active in Master Data',
                                    '', str(total_amount),
                                    str(user_data['gbp_rate']), current_user.username])
                else:
                    consolidated_writer.writerow(
                        [account_name, '', 'Account does not exist in Master Data or Marked as not processed', '',
                            str(total_amount),
                            str(user_data['gbp_rate']), current_user.username])
                    
            except Exception as e:
                    traceback.print_exc()
                    messages.error(request, f'!Error Occurred: {e}')
                    logging.basicConfig(filename=f"{output_dir}/Logger_{current_date_format}.log", level=logging.ERROR)
                    logging.error(account_name)
                    logging.error(e)

        for parent_comp in dic_group_comp:
            try:
                acc_list = dic_group_comp[parent_comp]
                pdfs = []
                send_to = ''
                send_cc = ''
                parent_comp_name = ''
                for x in acc_list:
                    if x['account_name'] == parent_comp:
                        send_to = x['email_ids_TO']
                        send_cc = x['email_ids_cc']
                        parent_comp_name = x['parent_comp_name']
                    elif send_to == '' and send_cc == '':
                        send_to = x['email_ids_TO']
                        send_cc = x['email_ids_cc']
                        parent_comp_name = x['parent_comp_name']
                for x in acc_list:
                    pdfs.append(x['pdf_file'])
                merger = PdfMerger()
                for pdf in pdfs:
                    merger.append(pdf)
                merger.write(f'{output_dir}/consolidated_{parent_comp_name}.pdf')
                merger.close()
                # Move files
                for pdf in pdfs:
                    file1 = os.path.basename(pdf)
                    shutil.move(
                        f'{output_dir}/{file1}',
                        f'{archive_dir}/{file1}'
                    )

                dest_wb = Workbook()
                filenames = []
                for x in acc_list:
                    filenames.append(x['excel_file'])

                i = 0
                for file in filenames:
                    i = i + 1
                    file1 = os.path.basename(file)
                    file_name = (os.path.splitext(file1)[0]).split(' ')[0] + str(i)
                    dest_wb.create_sheet(file_name)
                    dest_ws = dest_wb[file_name]
                    # Read source data
                    source_wb = load_workbook(file)
                    source_sheet = source_wb.active
                    for row in source_sheet.rows:
                        for cell in row:
                            dest_ws[cell.coordinate] = cell.value

                    # =================#
                del dest_wb['Sheet']
                for sheet in dest_wb:
                    sheet._legacy_drawing = None
                dest_wb.save(f'{output_dir}/consolidated_{parent_comp_name}.xlsx')
                for file in filenames:
                    file1 = os.path.basename(file)
                    shutil.move(
                        f'{output_dir}/{file1}',
                        f'{archive_dir}/{file1}'
                    )
                sleep(2)

                files = [f'{output_dir}/consolidated_{parent_comp_name}.pdf',
                            f'{output_dir}/consolidated_{parent_comp_name}.xlsx']
                if isSendEmail.lower() == 'true':
                    send_group_mail(
                        send_from=user_data['smtp_sender'],
                        send_to=send_to,
                        send_cc=send_cc,
                        files=files,
                        smtp_server=user_data['smtp_server'],
                        smtp_port=user_data['smtp_port'],
                        smtp_user=user_data['smtp_user'],
                        smtp_pass=user_data['smtp_pass'],
                        receiver_name=parent_comp_name
                    )
                    logging.basicConfig(
                        filename=f"{archive_dir}/Logger_{current_date_format}.log",
                        level=logging.INFO)
                    logging.info(str(parent_comp_name) + str(send_to+send_cc) + str(files))

                shutil.move(
                    f'{output_dir}/consolidated_{parent_comp_name}.pdf',
                    f'{archive_dir}/consolidated_{parent_comp_name}.pdf'
                )
                shutil.move(
                    f'{output_dir}/consolidated_{parent_comp_name}.xlsx',
                    f'{archive_dir}/consolidated_{parent_comp_name}.xlsx'
                )
            except Exception as e:
                sleep(1)
                logging.basicConfig(filename=f"{output_dir}/Logger_{current_date_format}.log", level=logging.ERROR)
                logging.error(parent_comp_name)
                logging.error('Group Company '+str(e))
    
    files = [f'{output_dir}/{Consolidated_Report}']
    if isSendEmail.lower() == 'true':
        send_consolidated_mail(
            send_from=user_data['smtp_sender'],
            send_to=ConsolidatedEmail,
            files=files,
            smtp_server=user_data['smtp_server'],
            smtp_port=user_data['smtp_port'],
            smtp_user=user_data['smtp_user'],
            smtp_pass=user_data['smtp_pass']

        )
        logging.basicConfig(
            filename=f"{archive_dir}/Logger_{current_date_format}.log",
            level=logging.INFO)
        logging.info(str(ConsolidatedEmail)+str(files))
    shutil.move(
        f'{output_dir}/{Consolidated_Report}',
        f'{archive_dir}/{Consolidated_Report}'
    )


#process_de_cp_desk function-------------------------------------------------------------------------------

def process_de_cp_desk_data(request, cp_desk_file, user_data):
    Word_To_PDF_OBJ = None
    current_user = User.objects.filter(id=user_data['user']).first()
    try:
        data = json.loads(cp_desk_file)
        df = pd.DataFrame(data)
        with io.BytesIO() as output:
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False)
            workbook = output.getvalue()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=output.xlsx'
        response.write(workbook)
    except Exception as e:
        messages.error(request, f"!Error Occurred: {str(e)}")
    messages.info(request, '!Checking File')

    processed_data = {}
    df = pd.read_excel(workbook)
    for index, row in df.iterrows():
        if pd.isna(row['Principals']) == False:
            if row['Principals'] in processed_data:
                processed_data[row['Principals']].append(row.to_dict())
            else:
                processed_data[row['Principals']] = [row.to_dict()]
    Word_To_PDF_OBJ = Word_To_PDF()
    
    isSendEmail = 'True'
    today = datetime.date.today()
    current_year_number = today.strftime("%Y")
    last_month_number = str(int(today.strftime("%m")) - 1).zfill(2)
    last_year_number = today.strftime("%Y")
    current_date_format = today.strftime("%d%m%Y")
    Consolidated_Report = f'CP_Desk_Consolidated_Report_{last_month_number}_{last_year_number}.csv'
    confirmation_exists = 'False'
    with open(f"{output_dir}/{Consolidated_Report}", 'w', newline='') as consolidated_write_file:
        consolidated_writer = csv.writer(consolidated_write_file, csv.unix_dialect)
        consolidated_writer.writerow(
            ['Account Name', 'Email_ids', 'Invoice_Number', 'User'])
        for account_name in processed_data:
            try:
                invoice_number = processed_data[account_name][0]['TLA']
                invoice_pdf_file = str(invoice_number)+'-CP'+str('-')+str(last_month_number)+str(last_year_number)+str('.pdf')
                invoice_excel_file = str(invoice_number)+'-CP'+str('-')+str(last_month_number)+str(last_year_number)+str('.xlsx')
                email_ids_TO = str(processed_data[account_name][0]['To']).split(';')
                cleanedList = str(processed_data[account_name][0]['CC']).split(';')

                if os.path.exists(f'{input_dir}/{invoice_pdf_file}') and os.path.exists(f'{input_dir}/{invoice_excel_file}'):
                    files = [f'{input_dir}/{invoice_pdf_file}', f'{input_dir}/{invoice_excel_file}']
                    invoice_exists = CpDeskInvoices.objects.filter(account_name=account_name, previous_month=last_month_number, previous_year=last_year_number).exists()
                    confirm_creating_invoice = True
                    if not invoice_exists == False:
                        if confirmation_exists.lower() == 'false':
                            confirm_creating_invoice = True
                            if confirm_creating_invoice:
                                confirmation_exists = 'true'

                    if confirm_creating_invoice:
                        messages.success(request, f"!Creating invoice for {account_name} and !Sending email to {account_name}")

                        please_confirm_text = "Please confirm we are able to deduct these lifting fees from the respective currency balance."
                        if isSendEmail.lower() == 'true':
                                            send_mail(
                                                send_from=user_data['smtp_sender'],
                                                send_to=email_ids_TO,
                                                send_cc=cleanedList,
                                                files=files,
                                                receiver_name=account_name,
                                                smtp_server=user_data['smtp_server'],
                                                smtp_port=user_data['smtp_port'],
                                                smtp_user=user_data['smtp_user'],
                                                smtp_pass=user_data['smtp_pass'],
                                    
                                            )
                        if not invoice_exists:
                            invoices = CpDeskInvoices.objects.create(
                                        account_name=account_name, 
                                        invoice_file_pdf=invoice_pdf_file,
                                        invoice_file_excel=invoice_excel_file, 
                                        invoice_number=invoice_number,
                                        previous_month=last_month_number, 
                                        previous_year=last_year_number,
                                        current_year=current_year_number,
                                        user_name=current_user.username,
                                    )
                        logging.basicConfig(
                            filename=f"{archive_dir}/Logger_{current_date_format}.log",
                            level=logging.INFO)
                        logging.info(
                            str(account_name) + str(str(email_ids_TO) + str(cleanedList)) + str(
                                files))
                        messages.success(request, f"!Email sent to{account_name}")
                    consolidated_writer.writerow(
                        [account_name, str(email_ids_TO)+str(cleanedList),invoice_number, current_user.username])
                else:
                    consolidated_writer.writerow(
                        [account_name, str(email_ids_TO) + str(cleanedList),'Invoice(s) do not exist in local folder',
                            current_user.username])
            except Exception as e:
                sleep(1)
                logging.basicConfig(filename=f"{output_dir}/Logger_{current_date_format}.log", level=logging.ERROR)
                logging.error(account_name)
                logging.error(e)

# WORD TO PDF---------------------------------------------------------------------------------------------

class Word_To_PDF:

    def __init__(self):
        super().__init__()
        # print(Invoice_Date)
        # self.invoice_date = Invoice_Date

        self.today = datetime.datetime.now()
        self.first = self.today.replace(day=1)
        self.last_month = self.first - datetime.timedelta(days=1)
        self.last_month_b = self.last_month.strftime('%b')
        self.current_month = self.today.strftime('%b')
        self.current_year = self.last_month.strftime('%y')  # today.strftime('%y')
        self.current_date = f"{self.today.strftime('%d')}-{self.current_month}-{self.today.strftime('%Y')}"

    def generate_pdf(self,account_name: str, address: str, unit_price: str, vat_type: str, gbp_rate: str,num_of_invoice: str):

        due_date = self.today + datetime.timedelta(days=7)
        due_date_str = f"{due_date.strftime('%d')}-{due_date.strftime('%b')}-{due_date.strftime('%Y')}"

        # invoice_number = f"{last_year_number}/{last_month_number}-{num_of_invoice}"
        invoice_number = f"{num_of_invoice}"

        unit_price = float(unit_price)  # convert to a float
        unit_price = round(unit_price, 2)  # round to 2 decimal places
        gbp_rate = float(gbp_rate) 
        unit_price_gbp = unit_price / gbp_rate
        unit_price_gbp = round(unit_price_gbp, 2)

        # invoice_title = f"{account_name}-Lifting Fees-{today.strftime('%d')}{today.strftime('%m')}{today.strftime('%Y')}"
        invoice_title = f"Invoice_{invoice_number}_{account_name}"

        newaddress = address.rsplit(',', 1)[0]
        Country = ''
        if len(address.rsplit(',', 1)) > 1:
            Country = address.rsplit(',', 1)[1]

        doc = DocxTemplate(f"template/Invoice_Template.docx")
        context = {
            'Address': newaddress,
            'InvoiceDate': self.current_date,
            'InvoiceNumber': invoice_number,
            'Reference': f"Lifting Fees - {self.last_month_b} {self.current_year}",
            'Unitprice': unit_price,
            'DueDate': due_date_str,
            'Amount': unit_price,
            'gbpRate':gbp_rate,
            'vatType':vat_type,
            'gbpAmount':unit_price_gbp,
            'ClientName': account_name,
            'Country': Country,
        }
        doc.render(context)
        doc.save(f"{output_dir}/{invoice_title}.docx")
        sys.stderr = open("consoleoutput.log", "w")
        pythoncom.CoInitialize()
        convert(f"{output_dir}/{invoice_title}.docx", f"{output_dir}/{invoice_title}.pdf")
        pythoncom.CoUninitialize()
        if os.path.exists(f"{output_dir}/{invoice_title}.docx"):
            os.remove(f"{output_dir}/{invoice_title}.docx")
        return [f"{invoice_title}.pdf", invoice_number]

    def genrate_excel(self,account_name: str, lifting_data: list):
        file_name = f"{account_name}-Lifting Fees-{self.today.strftime('%d')}{self.today.strftime('%m')}{self.today.strftime('%Y')}"

        workbook = xlsxwriter.Workbook(f'{output_dir}/{file_name}.xlsx')

        cell_fill = workbook.add_format()
        cell_fill.set_pattern(1)
        cell_fill.set_bg_color('#B4C6E7')
        cell_fill.set_bold()

        worksheet = workbook.add_worksheet()
        worksheet.set_column(0, 1, 35)
        worksheet.set_column(2, 27, 15)

        worksheet.write('A2', 'Account Name', cell_fill)
        worksheet.write('B2', account_name)

        worksheet.write('A3', f"Lifting Fee For {self.last_month_b} {self.current_year}")
        worksheet.write('A4', 'Classification : Private and Confidential')

        col_name_exists = False
        col_name_row = 4
        data_row = 5

        for data in lifting_data:
            col = 0

            for index in data:
                if not col_name_exists:
                    worksheet.write(col_name_row, col, index, cell_fill)

                try:
                    worksheet.write(data_row, col, data[index])
                except:
                    pass

                col += 1

            col_name_exists = True
            data_row += 1

        workbook.close()

        return f'{file_name}.xlsx'
    