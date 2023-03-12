import datetime
import os
import traceback
from django.db import IntegrityError
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.urls import reverse
from .forms import UserDataForm
import pandas as pd
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import openpyxl, json
from django.core import serializers
from .utils import process_de_data, process_de_cp_desk_data
from django.conf import settings
from time import sleep
from .models import (UserData,
                    EmailCondition,
                    GroupComp,
                    MasterData,
                    Invoices,
                    AccountList,
                    CpDeskInvoices,
                    )

type = 'normal'
today = datetime.datetime.now()
last_month_number = str(int(today.strftime("%m")) - 1).zfill(2)
current_year_number = today.strftime("%Y")
last_year_number = today.strftime("%Y")
CurrentDate = today.strftime("%d%m%Y")

output_dir = settings.MEDIA_ROOT[1]
archive_dir = settings.MEDIA_ROOT[2]

# Create your views here.
@login_required
def dashboard(request):
    template_admin = 'dashboard.html'
    template_user = 'user/user_dashboard.html'

    if request.user.is_authenticated:
        if request.user.is_admin:
            context = {
                'title': 'Dashboard'
            }
            return render(request, template_admin, context)
        else:
            context = {
                'title': 'Dashboard'
            }
            return render(request, template_user, context)
    else:
        return redirect('account/login')

@login_required
def settings(request):
    template = 'admin/settings.html'
    user_data = UserData.objects.filter(user_id=request.user.id).first()
    print(f"this is type of you getting from ORM  {type(user_data)}")
    context = {
        'title': 'Settings',
        'user_data' : user_data
    }

    return render(request, template, context)

@login_required
def upload_files(request):
    template = 'admin/upload-files.html'

    context = {
        'title': 'Upload Files'
    }

    return render(request, template, context)

@login_required
def user_page(request):

    template = 'admin/users.html'
    User = get_user_model()
    users = User.objects.filter(is_user=True)
    
    context = {
        "get_all_is_user": users,
        'title': 'Users'
    }
    return render(request, template, context)

@login_required
def cp_desk_view(request):
    template = 'user/cp_desk.html'

    context = {
        'title': 'Cp_Desk'
    }
    return render(request, template, context)

@login_required
def insert_user_data(request):
    user_name = request.user
    
    if request.method == 'POST':
        form = UserDataForm(request.POST)
        if form.is_valid():
            gbp_rate = form.cleaned_data['gbp_rate']
            smtp_server = form.cleaned_data['smtp_server']
            smtp_port = form.cleaned_data['smtp_port']
            smtp_user = form.cleaned_data['smtp_user']
            smtp_sender = form.cleaned_data['smtp_sender']
            smtp_pass = form.cleaned_data['smtp_pass']

            user = UserData.objects.filter(user=user_name).first()

            if user:
                user.gbp_rate = gbp_rate
                user.smtp_server = smtp_server
                user.smtp_port = smtp_port
                user.smtp_user = smtp_user
                user.smtp_sender = smtp_sender
                user.smtp_pass = smtp_pass
                user.save()
                messages.success(request, 'User Data Updated Successfully')
                return redirect('settings')

            else:
                UserData.objects.create(
                    gbp_rate=gbp_rate,
                    smtp_server=smtp_server,
                    smtp_port=smtp_port,
                    smtp_user=smtp_user,
                    smtp_sender=smtp_sender,
                    smtp_pass=smtp_pass,
                    user = request.user
                )

            messages.success(request, 'User Data Uploaded Successfully')
            return redirect('settings')

    else:
        form = UserDataForm()

    return render(request, 'settings.html', {'form': form})

def update_user_data(request, id):
    context ={}
 
    # fetch the object related to passed id
    obj = get_object_or_404(UserData, id = id)
 
    # pass the object as instance in form
    form = UserDataForm(request.POST or None, instance = obj)
 
    # save the data from the form and
    # redirect to detail_view
    if form.is_valid():
        form.save()
        messages.success(request, 'User Data Updated Successfully')
        return redirect("settings")
    # add form dictionary to context
    context["form"] = form
 
    return render(request, "update_view.html", context)

@login_required
def upload_master_data(request):
    # Open the excel file and get the active sheet
    try:
        if request.method == 'POST' and request.FILES['master_data_file']:
            file = request.FILES['master_data_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
    except:
        messages.error(request, "Invalid Excel File")
        return redirect('dashboard')

    # Validate the columns
    expected_columns = ["*ContactName", "Customer Address", "Email id", "CC email id", "*Description", "VAT Type", "Status"]
    header = [cell.value for cell in ws[1]]
    if header != expected_columns:
        messages.error(request, "Invalid Data")
        return redirect('dashboard')

    # Map the columns to the corresponding fields in the SQL table
    column_mapping = {
        "*ContactName": "contact_name",
        "Customer Address": "customer_address",
        "Email id": "email_id",
        "CC email id": "cc_email_id",
        "*Description": "description",
        "VAT Type": "vat_type",
        "Status": "status",
    }

    success_message_displayed = False

    # Iterate through the rows and add the data to the SQL table
    for row in ws.iter_rows(min_row=2):
        contact_name = row[0].value
        data = {
            column_mapping[header[i]]: row[i].value
            for i in range(1, len(header))
        }

        # Check if a row with the same contact_name and user already exists
        existing_data = MasterData.objects.filter(contact_name=contact_name, user=request.user).first()

        if existing_data:
            # Update the existing row with the new data
            existing_data.customer_address = data["customer_address"]
            existing_data.email_id = data["email_id"]
            existing_data.cc_email_id = data["cc_email_id"]
            existing_data.description = data["description"]
            existing_data.vat_type = data["vat_type"]
            existing_data.status = data["status"]
            existing_data.save()
            if not success_message_displayed:
                messages.success(request, 'Master Data Updated Successfully')
                success_message_displayed = True
        else:
            # Create a new row with the user's data
            master_data = MasterData.objects.create(
                contact_name=contact_name,
                customer_address=data["customer_address"],
                email_id=data["email_id"],
                cc_email_id=data["cc_email_id"],
                description=data["description"],
                vat_type=data["vat_type"],
                status=data["status"],
                user=request.user
            )
            if not success_message_displayed:
                messages.success(request, ('Master Data Uploaded Successfully'))
                success_message_displayed = True
    return redirect('dashboard')

@login_required
def add_email_condition(request):
    try:
        # Open the excel file and get the active sheet
        if request.method == 'POST' and request.FILES['email_data_file']:
            file = request.FILES['email_data_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
    except:
        messages.error(request, "'Invalid Excel File'")
        return redirect('dashboard')

    # Validate the columns
    expected_columns = ["Name", "Type"]
    header = [cell.value for cell in ws[1]]
    if 'Sr.No.' in header:
        header.remove('Sr.No.')
    if header != expected_columns:
        messages.error(request, "'Invalid Data'")
        return redirect('dashboard')

    # Map the columns to the corresponding fields in the SQL table
    column_mapping = {
        "Name": "contact_name",
        "Type": "type",
    }

    # Initialize the variable for success message display
    success_message_displayed = False

    # Iterate through the rows and add the data to the SQL table
    for row in ws.iter_rows(min_row=2):
        contact_name = row[1].value
        data = {
            column_mapping[header[i]]: row[i+1].value
            for i in range(len(header))
        }

        try:
            # Check if a row with the same contact_name and user already exists
            existing_data = EmailCondition.objects.get(contact_name=contact_name, user=request.user)

            # Update the existing row with the new data
            existing_data.type = data["type"]
            existing_data.save()

            if not success_message_displayed:
                messages.success(request, 'Email Data Updated Successfully')
                success_message_displayed = True
        except EmailCondition.DoesNotExist:
            # Create a new row with the user's data
            email_data = EmailCondition.objects.create(
                contact_name=contact_name,
                type=data["type"],
                user=request.user
            )

            if not success_message_displayed:
                messages.success(request, 'Email Data Uploaded Successfully')
                success_message_displayed = True
    return redirect('dashboard')

@login_required
def add_group_condition(request):
    
    try:
        # Open the excel file and get the active sheet
        if request.method == 'POST' and request.FILES['parent_child_file']:
            file = request.FILES['parent_child_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
    except:
        messages.error(request, "'Invalid Excel File'")
        return redirect('dashboard')
        
    # Validate the columns
    expected_columns = ["Parent", "Child","Parent_Company_Name"]
    header = [cell.value for cell in ws[1]]
    if header != expected_columns:
        messages.error(request, "'Invalid Data''")
        return redirect('dashboard')

    # Map the columns to the corresponding fields in the SQL table
    column_mapping = {
        "Parent": "Parent_Comp",
        "Child": "Child_Comp",
        "Parent_Company_Name" : "Parent_Company_Name"
    }
    
    # Initialize the variable for success message display
    success_message_displayed = False

    # Iterate through the rows and add the data to the SQL table
    for row in ws.iter_rows(min_row=2):
        Parent_Comp = row[0].value
        data = {
            column_mapping[header[i+1]]: row[i+1].value if row[i+1].value else None
            for i in range(len(header)-1)
        }

        if data["Child_Comp"] is not None:
            try:
                # Check if a row with the same Parent_Comp and Child_Comp already exists
                existing_data = GroupComp.objects.get(Parent_Comp=Parent_Comp, Child_Comp=data["Child_Comp"], user=request.user)

                # Update the existing row with the new data
                existing_data.Parent_Company_Name = data["Parent_Company_Name"]
                existing_data.save()
                if not success_message_displayed:
                    messages.success(request, 'Group Parent-child Data Updated Successfully')
                    success_message_displayed = True
                    
            except GroupComp.DoesNotExist:
                # Create a new row with the user's data
                try:
                    group_parent_child_data = GroupComp.objects.create(
                        Parent_Comp=Parent_Comp,
                        Child_Comp=data["Child_Comp"],
                        Parent_Company_Name=data["Parent_Company_Name"],
                        user=request.user
                    )
                except IntegrityError as e:
                    messages.error(request, 'Error: {}'.format(str(e)))
                    return redirect('dashboard')
                    
                if not success_message_displayed:
                    messages.success(request, 'Group Parent-child Data Uploaded Successfully')
                    success_message_displayed = True
                    
            except GroupComp.MultipleObjectsReturned:
                # Handle the case when multiple rows with the same Parent_Comp and Child_Comp values exist
                messages.error(request, 'Multiple rows with the same Parent_Comp and Child_Comp values exist')
                return redirect('dashboard')
                    
        else:
            # Delete the row with null values
            GroupComp.objects.filter(Parent_Comp=Parent_Comp, Child_Comp__isnull=True, user=request.user).delete()
            if not success_message_displayed:
                messages.error(request, 'Group Parent-child Data null row has been deleted')
                success_message_displayed = True
    return redirect('dashboard')

@login_required
def create_account_list(request):
    template_name = 'user/user_dashboard.html'
    try:
        # Open the excel file and get the active sheet
        if request.method == 'POST' and request.FILES['account_data_file']:
            file = request.FILES['account_data_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
    except:
        messages.error(request, "'Invalid Excel File'")
        return redirect('dashboard')

    expected_columns = ["Account Name", "isConsidered"]
    header = [cell.value for cell in ws[1]]
    header = [col_name for col_name in header if col_name != 'SrN0.']
    if header != expected_columns:
        messages.error(request, "Invalid Data")
        return redirect('dashboard')
    
    column_mapping = {
        "Account Name": "account_name",
        "isConsidered": "is_considered",
    }
    success_message_displayed = False

    for row in ws.iter_rows(min_row=2):
        account_name = row[1].value
        is_considered = row[2].value

        try:
            if is_considered == True:
                # Check if a row with the same contact_name and user already exists
                existing_data = AccountList.objects.get(account_name=account_name, user=request.user)

                # Update the existing row with the new data
                existing_data.is_considered = is_considered
                existing_data.save()

                if not success_message_displayed:
                    messages.success(request, 'Account_list Updated Successfully')
                    success_message_displayed = True
        except AccountList.DoesNotExist:
            # Create a new row with the user's data
            if is_considered == True:
                account = AccountList.objects.create(account_name=account_name, user=request.user, is_considered=True)
                account.save()
                if not success_message_displayed:
                    messages.success(request, 'Account_list uploaded successfully.')
                    success_message_displayed = True
    return redirect('dashboard')

@login_required
def browse_lifting_file(request):
    template_name = 'user/user_dashboard.html'
    confirm_continue_result = False

    if request.method == 'POST':
        if 'lift_fee_file' in request.FILES:
            file = request.FILES['lift_fee_file']

            try:
                wb = openpyxl.load_workbook(file)
            except:
                messages.error(request, "'Invalid Excel File'")
                return redirect('dashboard')
            

            # DefaultPath = settings.MEDIA_ROOT

            lift_fee_file = pd.read_excel(file)
            required_columns = ['Buy', 'Buy Amount', 'Sell', 'Sell Amount', 'Fee', 'Total Settlement', 'Beneficiary', 'When Booked', 'When Created', 'Created By', 'Delivery Method', 'Reference', 'Delivery Country ISO Code', 'Delivery Country Name', 'Your Reference', 'Cheque Number', 'Beneficiary ID', 'Execution Date', 'Payment Line Status', 'Payment Number', 'Processing Date', 'Bank Reference Number', 'AccountName', 'Bank Value Date', 'Exchange Rate', 'Our Reference', 'Charges Type', 'USD AMOUNT']

            
            if not set(required_columns).issubset(lift_fee_file.columns):
                messages.error(request, '!Invalid Lifting fee file.')
                return redirect('dashboard')

            user_data = UserData.objects.all()
            if len(user_data) == 0:
                messages.error(request, '!No Email configuration found, please contact with admin')
                return render(request, template_name)
            else:
                user_data = serializers.serialize('json', user_data)  # serialize the UserData object to JSON
                user_data = json.loads(user_data)[0]['fields']  # deserialize the JSON to a dictionary and get the fields

            lift_fee_data = lift_fee_file.to_dict(orient='records')

            for row in lift_fee_data:
                for key, value in row.items():
                    if isinstance(value, pd.Timestamp):
                        row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            lift_fee_json_data = json.dumps(lift_fee_data)
            request.session['lift_fee_json_data'] = lift_fee_json_data
            request.session['user_data'] = user_data

            template_data = {
                'lift_fee_json_data': lift_fee_json_data,
                'user_data': user_data,
                'confirm_url': reverse,
            }
            return render(request, 'user/confirm_continue.html', template_data)

        elif 'confirm_continue' in request.POST:
            confirm_continue_result = request.POST.get('confirm_continue')
            if confirm_continue_result == 'true':
                confirm_continue_result = True
            else:
                confirm_continue_result = False

            if not confirm_continue_result:
                return redirect('browse_lifting_file')

            lift_fee_json_data = request.session.get('lift_fee_json_data')
            user_data = request.session.get('user_data')

        try:
            process_de_data(request, lift_fee_json_data, user_data)
            messages.success(request, '!Process is completed successfully')

        except Exception as e:
            traceback.print_exc()
            messages.error(request, f'!Error Occurred: {e}')
        finally:
            messages.success(request, 'invoices are generated and sent over recipient email(s) address')

        return redirect('dashboard')
    else:
        messages.error(request, '!No file was uploaded.')
        return redirect('dashboard')

@login_required
def split_lifting_file(request):
    template_name = 'user/user_dashboard.html'

    try:
        # Open the excel file and get the active sheet
        if request.method == 'POST' and request.FILES['split_data_file']:
            file = request.FILES['split_data_file']
            wb = openpyxl.load_workbook(file)
    except:
        messages.error(request, "'Invalid Excel File'")
        return redirect('dashboard')
            # DefaultPath = settings.MEDIA_ROOT

    split_lift_fee_file = pd.read_excel(file)
    required_columns = ['Buy', 'Buy Amount', 'Sell', 'Sell Amount', 'Fee', 'Total Settlement', 'Beneficiary', 'When Booked', 'When Created', 'Created By', 'Delivery Method', 'Reference', 'Delivery Country ISO Code', 'Delivery Country Name', 'Your Reference', 'Cheque Number', 'Beneficiary ID', 'Execution Date', 'Payment Line Status', 'Payment Number', 'Processing Date', 'Bank Reference Number', 'AccountName', 'Bank Value Date', 'Exchange Rate', 'Our Reference', 'Charges Type', 'USD AMOUNT']

    
    if not set(required_columns).issubset(split_lift_fee_file.columns):
        messages.error(request, '!Invalid Lifting fee file.')
        return render(request, template_name)
    
    processed_data = {}
    df = pd.read_excel(file)

    for index, row in df.iterrows():
        if pd.isna(row['AccountName']) == False:
            if row['AccountName'] in processed_data:
                processed_data[row['AccountName']].append(row.to_dict())
            else:
                processed_data[row['AccountName']] = [row.to_dict()]

    messages.info(request, '!Generating invoices')
    
    

    processed_report_excelfile = f"{output_dir}/Processed_Report{CurrentDate}.xlsx"
    unprocessed_report_excelfile = f"{output_dir}/Unprocessed_Report{CurrentDate}.xlsx"
    group_report_excelfile = f"{output_dir}/Group_Report{CurrentDate}.xlsx"

    i = 0
    group_df = pd.DataFrame()
    processed_df = pd.DataFrame()
    unprocessed_df = pd.DataFrame()
    for account_name in processed_data:
        try:
            invoice_exists = Invoices.objects.filter(account_name=account_name, previous_month=last_month_number,
                                                    previous_year=last_year_number, invoice_Type=type).exists()
            is_exist_in_group = GroupComp.objects.filter(Child_Comp=account_name).exists()

            if is_exist_in_group:
                group_df = pd.concat([group_df, pd.DataFrame(processed_data[account_name])])
            elif invoice_exists:
                processed_df = pd.concat([processed_df, pd.DataFrame(processed_data[account_name])])
            else:
                unprocessed_df = pd.concat([unprocessed_df, pd.DataFrame(processed_data[account_name])])

        except Exception as e:
            traceback.print_exc()
            messages.error(request, f'!Error Occurred: {e}')

        if len(processed_df) > 0 :
            processed_df.to_excel(processed_report_excelfile,index=False)
        if len(unprocessed_df) > 0:
            unprocessed_df.to_excel(unprocessed_report_excelfile,index=False)
        if len(group_df) > 0:
            group_df.to_excel(group_report_excelfile, index=False)
    messages.success(request, "!processed_report and unprocessed_report is completed successfully")
    return redirect('dashboard')

@login_required
def get_Invoice_data(request):

    if os.path.exists(f'{output_dir}''\InvoiceData_'f'{last_month_number}''_'f'{last_year_number}''.xlsx'):
        os.remove(f'{output_dir}''\InvoiceData_'f'{last_month_number}''_'f'{last_year_number}''.xlsx')

    try:
        get_Invoice = Invoices.objects.filter(previous_month=last_month_number,
                                            previous_year=last_year_number, invoice_Type=type)

        if get_Invoice:
            df = pd.DataFrame(list(get_Invoice.values()))
            df.to_excel(f'{output_dir}/InvoiceData_{last_month_number}_{current_year_number}.xlsx', index=False)
            messages.success(request, "Invoice data successfully genrated")
            return redirect("dashboard")
        else:
            messages.error("Sorry not found any invoice data")
            return redirect("dashboard")

    except Exception as e:
        traceback.print_exc()
        messages.error(request, f'!Error Occurred: {e}')

@login_required
def browse_cp_desk_file(request):
    template_name = 'user/cp_desk.html'
    confirm_continue_result = False

    if request.method == 'POST':
        if 'cp_desk_data_file' in request.FILES:
            file = request.FILES['cp_desk_data_file']
            filename = file.name
            # DefaultPath = settings.MEDIA_ROOT

            try:
                wb = openpyxl.load_workbook(file)
            except:
                messages.error(request, "'Invalid Excel File'")
                return redirect('cp_desk_view')
            
            cp_desk_file = pd.read_excel(file)
            required_columns = ['TLA', 'Principals', 'Inv no', 'To', 'CC']

            if not set(required_columns).issubset(cp_desk_file.columns):
                messages.error(request, '!Invalid Cp_desk fee file.')
                return redirect('cp_desk_view')

            user_data = UserData.objects.all()
            if len(user_data) == 0:
                messages.error(request, '!No Email configuration found, please contact with admin')
                return render(request, template_name)
            else:
                user_data = serializers.serialize('json', user_data)  # serialize the UserData object to JSON
                user_data = json.loads(user_data)[0]['fields']  # deserialize the JSON to a dictionary and get the fields

            cp_desk_file = cp_desk_file.to_dict(orient='records')

            for row in cp_desk_file:
                for key, value in row.items():
                    if isinstance(value, pd.Timestamp):
                        row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            cp_desk_file = json.dumps(cp_desk_file)
        try:
            process_de_cp_desk_data(request, cp_desk_file, user_data)
            messages.success(request, '!Process is completed successfully')

        except Exception as e:
            traceback.print_exc()
            messages.error(request, f'!Error Occurred: {e}')
        finally:
            messages.success(request, 'invoices are generated and sent over recipient email(s) address')

        return redirect('cp_desk_view')
    else:
        messages.error(request, '!No file was uploaded.')
        return redirect('cp_desk_view')
    

@login_required
def split_cp_desk_file(request):
    template_name = 'user/cp_desk.html'
    try:
        # Open the excel file and get the active sheet
        if request.method == 'POST' and request.FILES['split_data_file']:
            file = request.FILES['split_data_file']
            wb = openpyxl.load_workbook(file)
    except:
        messages.error(request, "'Invalid Excel File'")
        return redirect('cp_desk_view')
            # DefaultPath = settings.MEDIA_ROOT

    split_cp_desk_file = pd.read_excel(file)
    required_columns = ['TLA', 'Principals', 'Inv no', 'To', 'CC']

    
    if not set(required_columns).issubset(split_cp_desk_file.columns):
        messages.error(request, '!Invalid Lifting fee file.')
        return render(request, template_name)
    
    processed_data = {}
    df = pd.read_excel(file)
    for index, row in df.iterrows():
        if pd.isna(row['Principals']) == False:
            if row['Principals'] in processed_data:
                processed_data[row['Principals']].append(row.to_dict())
            else:
                processed_data[row['Principals']] = [row.to_dict()]

    messages.info(request, '!Generating invoices')
    
    

    processed_report_excelfile = f"{output_dir}/CP_Desk_Processed_Report{CurrentDate}.xlsx"
    unprocessed_report_excelfile = f"{output_dir}/CP_Desk_Unprocessed_Report{CurrentDate}.xlsx"

    i = 0
    processed_df = pd.DataFrame()
    unprocessed_df = pd.DataFrame()
    for account_name in processed_data:
        i = i+1
        total_amount = 0
        group_parent_list = []
        try:
            invoice_exists = CpDeskInvoices.objects.filter(account_name=account_name, previous_month=last_month_number,
                                                    previous_year=last_year_number, invoice_Type=type).exists()

            if invoice_exists:
                processed_df = pd.concat([processed_df, pd.DataFrame(processed_data[account_name])])
            else:
                unprocessed_df = pd.concat([unprocessed_df, pd.DataFrame(processed_data[account_name])])

        except Exception as e:
            traceback.print_exc()
            messages.error(request, 'No data found in CpDeskInvoices for process and unprocess data')
            return redirect('cp_desk_view')

        if len(processed_df) > 0 :
            processed_df.to_excel(processed_report_excelfile,index=False)
        if len(unprocessed_df) > 0:
            unprocessed_df.to_excel(unprocessed_report_excelfile,index=False)
    messages.success(request, "!processed_report and unprocessed_report is completed successfully")
    return redirect('cp_desk_view')


@login_required
def get_cp_desk_Invoice_data(request):

    if os.path.exists(f'{output_dir}''\Cp_Desk_InvoiceData_'f'{last_month_number}''_'f'{last_year_number}''.xlsx'):
        os.remove(f'{output_dir}''\Cp_Desk_InvoiceData_'f'{last_month_number}''_'f'{last_year_number}''.xlsx')

    try:
        get_Invoice = CpDeskInvoices.objects.filter(previous_month=last_month_number,
                                            previous_year=last_year_number)

        if get_Invoice:
            df = pd.DataFrame(list(get_Invoice.values()))
            df.to_excel(f'{output_dir}/Cp_Desk_InvoiceData_{last_month_number}_{current_year_number}.xlsx', index=False)
            messages.success(request, "Cp_Desk_InvoiceData_ data successfully genrated")
            return redirect("cp_desk_view")
        else:
            messages.error("Sorry not found any Cp_Desk_InvoiceData")
            return redirect("cp_desk_view")

    except Exception as e:
        traceback.print_exc()
        messages.error("Sorry not found any Cp_Desk_InvoiceData")
        return redirect("cp_desk_view")